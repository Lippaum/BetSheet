import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, numbers
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
import base64

# Função para ajustar automaticamente a largura das colunas
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)  # Adiciona um pequeno padding
        ws.column_dimensions[column].width = adjusted_width

# Função para processar os dados (com cache)
@st.cache_data
def processar_dados(file, torneio, tip, winrate_min, winrate_max, data_inicio, data_fim, green_color, red_color):
    try:
        df = pd.read_excel(file, sheet_name="Tips Enviadas")
    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {str(e)}")
        return None, None
    
    # Conversão numérica
    df["Winrate 1"] = pd.to_numeric(df["Winrate 1"], errors="coerce")
    df["Winrate 2"] = pd.to_numeric(df["Winrate 2"], errors="coerce")
    df["Lucro/Prej."] = pd.to_numeric(df["Lucro/Prej."], errors="coerce")
    df["Lucro/Prej."] = df["Lucro/Prej."].fillna(0)
    # Assumindo que a coluna "J" contém as datas
    df["Data"] = pd.to_datetime(df.iloc[:, 9], errors="coerce")  # Índice 9 corresponde à coluna J (0-based)

    if "Campeonato" not in df.columns:
        st.error("A coluna 'Campeonato' não foi encontrada na planilha.")
        return None, None

    # Validação de winrate
    if not (0 <= winrate_min <= winrate_max <= 1):
        st.error("Winrates devem estar entre 0 e 100%, e o mínimo deve ser menor ou igual ao máximo")
        return None, None

    # Filtro dinâmico
    df_filtered = df
    if torneio != "Todos":
        df_filtered = df_filtered[df_filtered["Torneio"] == torneio]
    if tip != "Todos":
        df_filtered = df_filtered[df_filtered["Tip"].str.upper() == tip.upper()]
    df_filtered = df_filtered[
        (df_filtered["Winrate 1"] >= winrate_min) & 
        (df_filtered["Winrate 1"] <= winrate_max) &
        (df_filtered["Data"] >= pd.to_datetime(data_inicio)) &
        (df_filtered["Data"] <= pd.to_datetime(data_fim))
    ]

    if df_filtered.empty:
        st.warning("Nenhum dado disponível após o filtro.")
        return None, None

    # Normalizar confrontos
    df_filtered = df_filtered.copy()
    df_filtered["Confronto Normalizado"] = df_filtered.apply(
        lambda row: " vs ".join(sorted([str(row["Jogador A"]), str(row["Jogador B"])]))
        if pd.notna(row["Jogador A"]) and pd.notna(row["Jogador B"]) else "", axis=1
    )

    # Tabelas agrupadas
    df_confronto = df_filtered.groupby(["Torneio", "Confronto Normalizado"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_confronto["ROI (%)"] = (df_confronto["Lucro_Prej"] / df_confronto["Quantidade_Entradas"]).round(2)

    df_campeonato = df_filtered.groupby(["Torneio", "Campeonato"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_campeonato["ROI (%)"] = (df_campeonato["Lucro_Prej"] / df_campeonato["Quantidade_Entradas"]).round(2)

    # Adicionar tabelas ausentes
    df_winrate1 = df_filtered.groupby(["Torneio", "Winrate 1"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_winrate1["ROI (%)"] = (df_winrate1["Lucro_Prej"] / df_winrate1["Quantidade_Entradas"]).round(2)

    df_winrate2 = df_filtered.groupby(["Torneio", "Winrate 2"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_winrate2["ROI (%)"] = (df_winrate2["Lucro_Prej"] / df_winrate2["Quantidade_Entradas"]).round(2)

    # Criar confrontos de jogadores
    df_jogadores = df_filtered.melt(
        id_vars=["Torneio", "Lucro/Prej."],
        value_vars=["Jogador A", "Jogador B"],
        var_name="Posição",
        value_name="Jogador"
    )
    df_jogador = df_jogadores.groupby(["Torneio", "Jogador"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_jogador["ROI (%)"] = (df_jogador["Lucro_Prej"] / df_jogador["Quantidade_Entradas"]).round(2)

    # Criar confrontos de times
    df_times = df_filtered.melt(
        id_vars=["Torneio", "Lucro/Prej."],
        value_vars=["Time A", "Time B"],
        var_name="Posição",
        value_name="Time"
    )
    df_time = df_times.groupby(["Torneio", "Time"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_time["ROI (%)"] = (df_time["Lucro_Prej"] / df_time["Quantidade_Entradas"]).round(2)

    df_confronto_times = df_filtered.groupby(["Torneio", "Confronto Normalizado"]).agg(
        Quantidade_Entradas=("Lucro/Prej.", "count"),
        Lucro_Prej=("Lucro/Prej.", "sum")
    ).reset_index()
    df_confronto_times["ROI (%)"] = (df_confronto_times["Lucro_Prej"] / df_confronto_times["Quantidade_Entradas"]).round(2)

    # Salvar em Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_filtered.to_excel(writer, sheet_name="Tips Enviadas", index=False)
        df_campeonato.to_excel(writer, sheet_name="Campeonato", index=False)
        df_confronto.to_excel(writer, sheet_name="Confronto", index=False)
        df_winrate1.to_excel(writer, sheet_name="Winrate 1", index=False)
        df_winrate2.to_excel(writer, sheet_name="Winrate 2", index=False)
        df_jogador.to_excel(writer, sheet_name="Jogador", index=False)
        df_time.to_excel(writer, sheet_name="ROI por Time", index=False)
        df_confronto_times.to_excel(writer, sheet_name="Confronto Times", index=False)

    output.seek(0)
    wb = load_workbook(output)
    sheets = ["Tips Enviadas", "Campeonato", "Confronto", "Winrate 1", "Winrate 2", 
              "Jogador", "Time", "Confronto Times"]

    for sheet_name in sheets:
        ws = wb[sheet_name]
        if ws.max_row > 1 and ws.max_column > 1:
            table_range = f"A1:{ws.cell(row=ws.max_row, column=ws.max_column).coordinate}"
            table = Table(displayName=f"Table_{sheet_name.replace(' ', '_')}", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
            ws.add_table(table)

            # Cabeçalhos brancos e em negrito
            header_fill = PatternFill(start_color="B2B2B2", end_color="B2B2B2", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws[1]:  # Linha 1 (cabeçalho)
                cell.fill = header_fill
                cell.font = header_font

            # Ajuste automático das colunas
            auto_adjust_column_width(ws)

            # Formatação condicional para Lucro/Prej. e ROI (%)
            lucro_col_idx = None
            roi_col_idx = None
            torneio_col_idx = None
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value in ["Lucro/Prej.", "Lucro_Prej"]:
                    lucro_col_idx = col_idx
                elif cell.value == "ROI (%)":
                    roi_col_idx = col_idx
                elif cell.value == "Torneio":
                    torneio_col_idx = col_idx

            # Negrito em todas as linhas da coluna "Torneio"
            if torneio_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=torneio_col_idx, max_col=torneio_col_idx):
                    for cell in row:
                        cell.font = Font(bold=True)

            # Formatação numérica para "Lucro/Prej." com duas casas decimais
            if lucro_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=lucro_col_idx, max_col=lucro_col_idx):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                value = float(cell.value)
                                cell.value = round(value, 2)  # Arredondar para duas casas decimais
                                cell.number_format = '0.00'  # Formato com duas casas decimais
                                if value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
                                    cell.font = Font(bold=True, color="006400")  # Verde escuro em negrito
                                elif value < 0:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Vermelho claro
                                    cell.font = Font(bold=True, color="8B0000")  # Vermelho escuro em negrito
                            except (ValueError, TypeError):
                                continue

            # Formatação de "ROI (%)" como porcentagem
            if roi_col_idx:
                for row in ws.iter_rows(min_row=2, min_col=roi_col_idx, max_col=roi_col_idx):
                    for cell in row:
                        if cell.value is not None:
                            try:
                                value = float(cell.value)  # Valor como decimal (ex.: 0.25 para 25%)
                                cell.number_format = '0.00%'  # Formato de porcentagem com duas casas decimais
                                if value > 0:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
                                    cell.font = Font(bold=True, color="006400")  # Verde escuro em negrito
                                elif value < 0:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Vermelho claro
                                    cell.font = Font(bold=True, color="8B0000")  # Vermelho escuro em negrito
                            except (ValueError, TypeError):
                                continue

           # Formatação condicional específica para "Tips Enviadas" com base em "Green" ou "Red"
            if sheet_name == "Tips Enviadas":
                resultado_col_idx = None
                for col_idx, cell in enumerate(ws[1], start=1):
                    if cell.value == "Resultado":  # Assumindo que a coluna de resultado contém "Green" ou "Red"
                        resultado_col_idx = col_idx
                if resultado_col_idx:
                    for row in ws.iter_rows(min_row=2, min_col=resultado_col_idx, max_col=resultado_col_idx):
                        for cell in row:
                            if cell.value is not None:
                                if isinstance(cell.value, str) and "Green" in cell.value:
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Verde claro
                                    cell.font = Font(bold=True, color="006400")  # Verde escuro em negrito
                                elif isinstance(cell.value, str) and "Red" in cell.value:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Vermelho claro
                                    cell.font = Font(bold=True, color="8B0000")  # Vermelho escuro em negrito

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return df_filtered, output

    return df_filtered, output

# Interface principal
st.title("Análise Planilhas TipManager 📊")

uploaded_file = st.file_uploader("Arraste e solte o arquivo aqui", type=["xlsx"])

if uploaded_file:
    # Carregar dados para determinar a primeira data e o menor winrate
    df = pd.read_excel(uploaded_file, sheet_name="Tips Enviadas")
    primeira_data = df.iloc[:, 9].min()  # Coluna J (índice 9)
    if pd.isna(primeira_data):
        primeira_data = datetime(2023, 1, 1).date()  # Valor padrão se não houver dados
    else:
        primeira_data = pd.to_datetime(primeira_data).date()

    # Determinar o menor Winrate 1 válido
    df["Winrate 1"] = pd.to_numeric(df["Winrate 1"], errors="coerce")
    min_winrate = df["Winrate 1"].min()
    if pd.isna(min_winrate) or min_winrate < 0:
        min_winrate = 0  # Valor padrão se não houver winrate válido
    else:
        min_winrate = max(0, min_winrate)  # Garante que não seja negativo
    initial_winrate_min = int(min_winrate)  # Valor inicial do slider

    # Filtros na barra lateral
    with st.sidebar:
        st.header("Filtros")
        torneios = df["Torneio"].unique().tolist()
        torneios.insert(0, "Todos")
        torneio = st.selectbox("Torneio", options=torneios, index=0)
        
        # Ajustar os valores de "Tip" para incluir "Over" e "Under"
        tips = df["Tip"].str.upper().unique().tolist()
        tips = [t.capitalize() if t in ["over", "under"] else t for t in tips]  # Capitaliza "over" e "under"
        tips.insert(0, "Todos")
        tip = st.selectbox("Tip", options=tips, index=0)
        
        # Slider para intervalo de winrate com valor inicial baseado no menor winrate
        winrate_range = st.slider("Intervalo de Winrate (%)", 0, 100, (initial_winrate_min, 100), step=1)
        winrate_min, winrate_max = winrate_range[0] / 100, winrate_range[1] / 100  # Correção aplicada

        # Filtros de data ajustados
        data_inicio = st.date_input("Data Início", value=primeira_data)
        data_fim = st.date_input("Data Fim", value=datetime.now().date())

        st.header("Formatação")
        col3, col4 = st.columns(2)
        with col3:
            green_color = st.color_picker("Cor Positivo", "#90EE90")  # Verde claro como padrão
        with col4:
            red_color = st.color_picker("Cor Negativo", "#FF6347")   # Vermelho claro como padrão

    if winrate_min > winrate_max:
        st.error("Winrate Mínimo não pode ser maior que o Máximo.")
    elif data_inicio > data_fim:
        st.error("Data de Início não pode ser maior que a Data de Fim.")
    else:
        # Processar dados com spinner
        with st.spinner('Processando dados...'):
            result = processar_dados(uploaded_file, torneio, tip, winrate_min, winrate_max, 
                                   data_inicio, data_fim, green_color, red_color)
        if result[0] is not None:
            df_filtered, excel_file = result

            # Resumo em Cards customizados
            st.subheader("Resumo Geral")
            total_entradas = df_filtered.shape[0]
            lucro_total = df_filtered['Lucro/Prej.'].sum()
            roi_medio = (lucro_total / total_entradas * 100) if total_entradas > 0 else 0

            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown(
                    f'<div class="custom-metric"><span class="custom-metric-label">Total de Entradas</span><br><span class="custom-metric-value">{total_entradas}</span></div>',
                    unsafe_allow_html=True
                )
            with col2:
                st.markdown(
                    f'<div class="custom-metric"><span class="custom-metric-label">Lucro Total</span><br><span class="custom-metric-value">{lucro_total:.2f}</span><br><span class="custom-metric-indicator">{"✅" if lucro_total >= 0 else "❌"}</span></div>',
                    unsafe_allow_html=True
                )
            with col3:
                st.markdown(
                    f'<div class="custom-metric"><span class="custom-metric-label">ROI Médio</span><br><span class="custom-metric-value">{roi_medio:.2f}%</span><br><span class="custom-metric-indicator">{"✅" if roi_medio >= 0 else "❌"}</span></div>',
                    unsafe_allow_html=True
                )

            # Análise por torneio em tabela (aparece apenas se "Torneio" for "Todos")
            if torneio == "Todos":
                st.subheader("Análise por Torneio")
                df_torneio = df_filtered.groupby("Torneio").agg(
                    Total_Entradas=("Lucro/Prej.", "count"),
                    Lucro_Total=("Lucro/Prej.", "sum")
                ).reset_index()
                df_torneio["ROI (%)"] = ((df_torneio["Lucro_Total"] / df_torneio["Total_Entradas"]) * 100).round(2)
                # Formatar para duas casas decimais com ponto
                df_torneio["Lucro_Total"] = df_torneio["Lucro_Total"].apply(lambda x: f"{x:.2f}")
                df_torneio["ROI (%)"] = df_torneio["ROI (%)"].apply(lambda x: f"{x:.2f}%")  # Adicionado % ao ROI
                st.table(df_torneio[["Torneio", "Total_Entradas", "Lucro_Total", "ROI (%)"]])

            # Botão de download centralizado em um contêiner
            st.markdown('<div class="button-container">', unsafe_allow_html=True)
            st.download_button(
                label="Download Planilha Ajustada",
                data=excel_file,
                file_name="Analise_TipManager.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_button",
                help="Clique para baixar a planilha ajustada"
            )
            st.markdown('</div>', unsafe_allow_html=True)

# Removido "Desenvolvido com Streamlit"
